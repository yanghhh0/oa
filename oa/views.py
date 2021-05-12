import json

from django.utils.six import BytesIO
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
import requests
import qrcode
import time
import hashlib
from geopy.distance import geodesic
from openpyxl import load_workbook
from oa.api import is_login, check_cookie, check_login, model_to_dict, check_stu_login, \
    check_stu_login_cookie, url_change, search_items, student_check, get_check_record_table, \
    items_to_table, update_item, search_stu_item, add_stu_to_item, generate_check_info, get_check_info_api, \
    api_init
from django.views.decorators.clickjacking import xframe_options_sameorigin


LOCAL_HOST = '192.168.128.143'
item_to_name = {'name': '姓名', 'phone': '手机号', 'sex': '性别', 'age': '年龄', 'stu_number': '学号', 'college': '学院'
         , 'major': '专业', 'cls': '班级'}


@csrf_exempt
def login(request):
    dt = dict()
    if request.POST:
        info = request.POST
        print(info)
        dt['email'] = info.get('email')
        dt['password'] = info.get('password')
        m = hashlib.sha1()
        m.update(dt['password'].encode(encoding="utf-8"))
        m.hexdigest()
        dt['password'] = m.hexdigest()
        if check_login(dt['email'], dt['password']) == 0:
            # return render(request, 'teacher.html', dt)
            response = redirect('/teacher/')
            response.set_cookie('email', dt['email'], 7200)
            response.set_cookie('password', dt['password'], 7200)
            return response
        else:
            if check_login(dt['email'], dt['password']) == 2:
                dt['error_msg'] = '邮箱错误'
            else:
                dt['error_msg'] = '密码错误'
            return render(request, 'login.html', dt)
    else:
        flag, rank = check_cookie(request)
        if flag:
            return redirect('/teacher/')
        return render(request, 'login.html', {'error_msg': ''})


@csrf_exempt
def register(request):
    dt = dict()
    if request.POST:
        dt['email'] = request.POST['email']
        dt['password'] = request.POST['password']
        return render(request, 'login.html', dt)
    else:
        return render(request, 'register.html')


def teacher(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html')
    rank = model_to_dict(rank)
    print(json.dumps(rank))
    return render(request, 'teacher.html', rank)


@xframe_options_sameorigin
def check(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    keys = request.GET.keys()
    dt = dict()
    for key in keys:
        dt[key] = request.GET[key]
    print(dt)
    dt['uid'] = uid
    dt['time'] = round(time.time())
    # 生成签到记录
    dt['check_id'] = generate_check_info(t_uid=uid, check_time=dt['time'], item_id=dt['class_list'])
    if not dt.get('N'):
        dt['N'] = 181.0
    if not dt.get('E'):
        dt['E'] = 181.0
    print(dt)
    return render(request, 'check.html', dt)


def get_check_info(request):
    check_id = request.GET.get('check_id')
    res = get_check_info_api(check_id)
    res = json.dumps(res)
    return HttpResponse(res)


def logout(request):
    req = redirect('/login/')
    req.delete_cookie('email')
    req.delete_cookie('password')
    return req


def make_qrcode(request):
    print(request)
    dt = request.GET.dict()
    if dt.get('api'):
        url = f"http://{LOCAL_HOST}:8000/{dt.get('api')}/?"
    else:
        url = f"http://{LOCAL_HOST}:8000/stuCheck/?"
    url = url_change(url, dt)
    print(f'qrcode:{url}')
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image()
    buf = BytesIO()
    img.save(buf)
    img_stream = buf.getvalue()
    response = HttpResponse(img_stream, content_type="image/jpg")
    return response


def stu_login(request):
    dt = dict()
    if request.POST:
        info = request.POST
        dt['uid'] = info.get('uid')
        dt['password'] = info.get('password')
        # TODO：password 加密
        if check_stu_login(dt['uid'], dt['password']) == 0:
            # return render(request, 'teacher.html', dt)
            response = redirect('/stuIndex/')
            response.set_cookie('uid', dt['uid'], 7200)
            response.set_cookie('password', dt['password'], 7200)
            return response
        else:
            if check_stu_login(dt['uid'], dt['password']) == 2:
                dt['error_msg'] = '学号错误'
            else:
                dt['error_msg'] = '密码错误'
            return render(request, 'stu_login.html', dt)
    else:
        flag, rank = check_stu_login_cookie(request)
        if flag:
            return redirect('/stuIndex/')
        return render(request, 'stu_login.html', {'error_msg': ''})


def stu_index(request):
    # 如果当前时间和二维码的时间差距超过15s就当过期处理
    flag, rank = check_stu_login_cookie(request)
    if not flag:
        return render(request, 'stu_login.html')
    print(request)
    if not request.GET.get('time'):
        return render(request, 'stu_index.html', {'res_msg': '登录后请重新扫描二维码'})
    rank = model_to_dict(rank)
    check_time = int(request.GET.get('time'))
    t = time.time()
    t = int(round(t * 1000))
    print(f't:{t}, check_time:{check_time}')
    if t - check_time > 15 * 1000:
        return render(request, 'stu_index.html', {'res_msg': '二维码已过期'})
    res = student_check(rank['uid'], request.GET.get('cls_id'), request.GET.get('check_id'))
    return render(request, 'stu_index.html', {'res_msg': res})


def stu_check(request):
    if request.POST:
        uid = request.POST.get('stu_id')
        N = float(request.POST.get('N'))
        E = float(request.POST.get('E'))
        NN = float(request.GET.get('N'))
        EE = float(request.GET.get('E'))
        print(f'N={N}, E={E}, NN={NN}, EE={EE}')
        if float(NN) < 90 and float(EE) < 180:
            print(geodesic((N, E), (NN, EE)))
        check_time = int(request.GET.get('time'))
        t = time.time()
        t = int(round(t * 1000))
        if t - check_time > 15 * 1000:
            return render(request, 'stu_index.html', {'res_msg': '二维码已过期'})
        res = student_check(uid, request.GET.get('cls_id'), request.GET.get('check_id'))
        return render(request, 'stu_index.html', {'res_msg': res})
    dt = dict()
    dt['dt'] = json.dumps(dict(request.GET))
    print(dt['dt'])
    return render(request, 'stu_check.html', dt)


@xframe_options_sameorigin
def check_setting(request):
    dt = dict()
    print(request.GET)
    items = search_items(request.GET['uid'])
    dt['class_list'] = items
    print(dt)
    return render(request, 'check_setting.html', dt)


@xframe_options_sameorigin
def check_history(request):
    return render(request, 'check_history.html')


def get_record_table(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    total, table = get_check_record_table(uid, page=int(request.GET.get('page'))-1, limit=int(request.GET.get('limit')))
    for obj in table:
        obj['check_time'] = time.strftime("%Y年%m月%d日 %H：%M：%S", time.localtime(obj['check_time']))

    # test_data = {"check_id": 1, "check_item": "计算机171", "check_time": "2021/4/17/15:30", "checked_count": 20,
    #              "total_count": 21, "unchecked_stu": ['石头人']}
    # table.append(test_data)
    res = {"code": 0, "msg": "", "count": total, "data": table}
    res = json.dumps(res)
    return HttpResponse(res)


@xframe_options_sameorigin
def add_item(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    if request.POST:
        print(request.POST)
        update_item(uid, request.POST.get('name'), 0)
    return render(request, 'search_item.html')


def get_items(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    items = search_items(uid)
    items = items_to_table(items)
    res = {"code": 0, "msg": "", "count": len(items), "data": items}
    res = json.dumps(res)
    return HttpResponse(res)


@xframe_options_sameorigin
def search_item(request):
    return render(request, 'search_item.html')


def change_stu_item(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    if request.POST:
        print(1)
        # update_item(uid, request.POST.get('name'), 2, uid=request.POST.get('uid'))
    return render(request, 'search_item.html')


def del_item(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    uid = model_to_dict(rank).get('uid')
    if request.POST:
        update_item(uid, request.POST.get('name'), 1, request.POST.get('uid'))
    return render(request, 'search_item.html')


@xframe_options_sameorigin
def add_stu(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    dt = dict()

    if request.POST:
        dt['itemId'] = request.GET.get('itemId')
        print(request.POST)
        add_stu_to_item(dt['itemId'], request.POST.get('uid'))
    return render(request, 'add_stu.html', dt)


def import_item(request):
    file = request.FILES.get('file')
    wb = load_workbook(file)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])
    data_list = []
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = ws.cell(row=rx, column=2).value
        data_list.append({'id': w1, 'name': w2})
    # data_list = list(set(data_list))
    print(data_list)
    for obj in data_list:
        add_stu_to_item(request.GET.get('item_id'), obj['id'])
    res = {"code": 0, "msg": "", "res": 1}
    res = json.dumps(res)
    return HttpResponse(res)


@xframe_options_sameorigin
def search_stu(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    dt = dict()
    dt['itemId'] = request.GET.get('itemId')
    return render(request, 'add_stu.html', dt)


def get_stu_item(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    dt = search_stu_item(request.GET.get('itemId'))
    res = {"code": 0, "msg": "", "count": len(dt), "data": dt}
    res = json.dumps(res)
    return HttpResponse(res)


def test(request):
    return render(request, 'test.html')


@xframe_options_sameorigin
def index(request):
    return render(request, 'index.html')


@xframe_options_sameorigin
def recheck(request):
    flag, rank = check_cookie(request)
    if not flag:
        return render(request, 'login.html', {'error_msg': ''})
    dt = dict()
    dt['check_id'] = request.GET.get('check_id')
    return render(request, 'recheck.html', dt)


def stu_recheck(request):
    if request.POST:
        print(request.POST)
        print(request.GET)
        print(request)
        check_id = request.GET.get('check_id')
        cls_id = get_check_info_api(check_id).get('item_id')
        res_msg = student_check(request.POST.get('stu_id'), cls_id,
                                check_id, check_type=2,
                                check_reason=request.POST.get('recheck_reason'))
        return render(request, 'stu_index.html', {"res_msg": res_msg})

    dt = dict()
    dt['cls_id'] = request.GET.get('cls_id')
    dt['check_id'] = request.GET.get('check_id')
    return render(request, 'stu_recheck.html', dt)


# ----------------------------------------meeting_check--------------------------------------
@xframe_options_sameorigin
def meeting_check_setting(request):
    url = ""
    if request.POST:
        url = url_change('/mc_qrcode/?', request.POST)
    return render(request, 'meeting_check_setting.html', {"src": url})


def make_mc_qrcode(request):
    dt = request.GET.dict()
    url = f"http://{LOCAL_HOST}:8000/meetingCheck/?"
    url = url_change(url, dt)
    qr = qrcode.QRCode(box_size=12, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image()
    buf = BytesIO()
    img.save(buf)
    img_stream = buf.getvalue()
    return HttpResponse(img_stream, content_type="image/jpg")


def meeting_check(request):
    if request.POST:
        print(request.POST)
        return render(request, 'stu_index.html', {'res_msg': '签到成功'})
    dt = dict()
    lt = list(dict(request.GET).keys())
    dt['item'] = []
    for i in lt:
        dt['item'].append({'key': i, 'value': item_to_name.get(i)})
    print(dt)
    return render(request, 'meeting_check.html', dt)
# ----------------------------------------meeting_check--------------------------------------


# 初始化整个系统
def init(request):
    api_init()
    return render(request, 'stu_index.html', {"res_msg": 'ok!'})
